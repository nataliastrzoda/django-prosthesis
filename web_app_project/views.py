import csv
import io
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .models import *
from .forms import *
from django.core.paginator import Paginator
from django.db.models import Q, Count


def home(request):
    from .models import Patient, Prosthesis, Company, PatientProsthesis

    total_patients = Patient.objects.count()
    companies = Company.objects.all()
    
    # --- 1. Obliczenia dla kafelka: % pacjentów z co najmniej jednym 100% dopasowaniem ---
    patients_with_match = Patient.objects.filter(
        patientprosthesis__match_score__gte=100
    ).distinct().count()
    
    success_rate = 0
    if total_patients > 0:
        success_rate = (patients_with_match / total_patients) * 100

    # --- 2. Dane do wykresów ---
    chart_labels = [c.name for c in companies]
    
    # Wykres 1: Ilość protez w katalogu wg firm
    chart_data = [Prosthesis.objects.filter(company=c).count() for c in companies]
    
    # Wykres 2: % pacjentów, którzy znaleźli dopasowanie 100% w danej firmie
    success_chart_data = []
    for c in companies:
        c_match_count = Patient.objects.filter(
            patientprosthesis__match_score__gte=100,
            patientprosthesis__prosthesis__company=c
        ).distinct().count()
        
        c_percent = (c_match_count / total_patients * 100) if total_patients > 0 else 0
        success_chart_data.append(round(c_percent, 1))

    return render(request, "home.html", {
        "patient_count":    total_patients,
        "prosthesis_count": Prosthesis.objects.count(),
        "company_count":    companies.count(),
        "success_rate":     success_rate,
        "chart_labels":     chart_labels,
        "chart_data":       chart_data,
        "success_chart_data": success_chart_data,
    })


def prosthesis_list(request):
    prostheses = Prosthesis.objects.select_related("company").all().order_by("name")
    search = request.GET.get("search", "")

    if search:
        prostheses = prostheses.filter(
            Q(name__icontains=search) |
            Q(company__name__icontains=search)
        )

    min_price = request.GET.get("min_price", "")
    max_price = request.GET.get("max_price", "")

    if min_price:
        try:
            prostheses = prostheses.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            prostheses = prostheses.filter(price__lte=float(max_price))
        except ValueError:
            pass

    per_page = request.GET.get("per_page", 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    paginator = Paginator(prostheses, per_page)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    return render(request, "prosthesis_list.html", {
        "page_obj": page_obj,
        "search": search,
        "min_price": min_price,
        "max_price": max_price,
        "per_page": per_page,
    })


def patient_profile(request, id):
    patient = get_object_or_404(Patient, id=id)
    history = PatientProsthesis.objects.filter(patient=patient).select_related('prosthesis', 'prosthesis__company')

    # Filtrowanie po parametrach GET
    search_name = request.GET.get("search_name", "")
    search_company = request.GET.get("search_company", "")
    min_price = request.GET.get("min_price", "")
    max_price = request.GET.get("max_price", "")
    sort_by = request.GET.get("sort_by", "-match_score")

    if search_name:
        history = history.filter(prosthesis__name__icontains=search_name)
    if search_company:
        history = history.filter(prosthesis__company__name__icontains=search_company)
    if min_price:
        try: history = history.filter(prosthesis__price__gte=float(min_price))
        except ValueError: pass
    if max_price:
        try: history = history.filter(prosthesis__price__lte=float(max_price))
        except ValueError: pass

    # Sortowanie domyślne: od najlepszych dopasowań (zielone -> żółte -> czerwone)
    if sort_by in ["match_score", "-match_score", "prosthesis__price", "-prosthesis__price"]:
        history = history.order_by(sort_by)
    else:
        history = history.order_by('-match_score')

    return render(
        request, "patient_profile.html", 
        {
            "patient": patient, 
            "history": history,
            "search_name": search_name,
            "search_company": search_company,
            "min_price": min_price,
            "max_price": max_price,
            "sort_by": sort_by,
        }
    )

def matches_list(request):
    """Zupełnie nowy widok: Katalog dopasowań pacjentów."""
    # Adnotacja: Policz tylko te protezy, które pasują fizycznie (match_score >= 0)
    patients = Patient.objects.annotate(
        valid_matches=Count('patientprosthesis', filter=Q(patientprosthesis__match_score__gte=0))
    ).order_by('-valid_matches', 'last_name')

    search = request.GET.get("search", "")
    if search:
        patients = patients.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
        )

    per_page = request.GET.get("per_page", 10)
    try: per_page = int(per_page)
    except ValueError: per_page = 10

    paginator = Paginator(patients, per_page)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    return render(request, "matches_list.html", {
        "page_obj": page_obj,
        "search": search,
        "per_page": per_page,
    })


def patients(request):
    patients = Patient.objects.all()
    search = request.GET.get("search", "")
    min_budget = request.GET.get("min_budget", "")

    if search:
        patients = patients.filter(
            first_name__icontains=search
        ) | patients.filter(
            last_name__icontains=search
        )

    if min_budget:
        patients = patients.filter(budget__gte=min_budget)

    per_page = request.GET.get("per_page", 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    paginator = Paginator(patients.order_by("last_name"), per_page)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "patients.html",
        {
            "page_obj": page_obj,
            "search": search,
            "min_budget": min_budget,
            "per_page": per_page,
        },
    )


# ── LOGIKA ALGORYTMÓW DOPASOWANIA ─────────────────────────────────────────────

def calculate_size(level, circumference):
    """Logika przeliczająca obwód kikuta na rozmiar leja protetycznego."""
    if not circumference or not level:
        return None
    
    c = float(circumference)
    if level in ['Udo', 'Staw biodrowy']:
        if c < 40: return 'S'
        elif c <= 50: return 'M'
        else: return 'L'
    elif level == 'Podudzie':
        if c < 30: return 'S'
        elif c <= 40: return 'M'
        else: return 'L'
    elif level == 'Ramię':
        if c < 25: return 'S'
        elif c <= 32: return 'M'
        else: return 'L'
    elif level in ['Przedramię', 'Dłoń']:
        if c < 20: return 'S'
        elif c <= 26: return 'M'
        else: return 'L'
    return 'M'

def calculate_match_score(patient, prosthesis):
    """Logika obliczająca wynik dopasowania protezy do pacjenta."""
    # 1. Filtr bezwzględny (Twardy)
    if (patient.limb_type != prosthesis.limb_type or
        patient.side != prosthesis.side or
        patient.amputation_level != prosthesis.amputation_level or
        patient.size != prosthesis.size):
        return -1.0  # -1 oznacza brak dopasowania (czerwony X)
    
    # 2. Filtr budżetowy (Miękki)
    if prosthesis.price <= 0:
        return 100.0  # Zabezpieczenie przed błędem dzielenia przez zero
        
    percentage = (patient.budget / prosthesis.price) * 100
    return min(percentage, 100.0)  # Maksymalnie 100%


def add_patient(request):
    if request.method == "POST":
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.size = calculate_size(patient.amputation_level, patient.circumference)
            patient.save()
            
            # Automatyczne dopasowanie do wszystkich protez w bazie
            for prosthesis in Prosthesis.objects.all():
                score = calculate_match_score(patient, prosthesis)
                PatientProsthesis.objects.create(patient=patient, prosthesis=prosthesis, match_score=score)
            
            messages.success(request, f"Pacjent {patient.first_name} {patient.last_name} zapisany! System wygenerował dopasowania (Rozmiar: {patient.size}).")
            return redirect("/")
    else:
        form = PatientForm()
    return render(request, "add_patient.html", {"form": form})


def add_prosthesis(request):
    if request.method == "POST":
        form = ProsthesisForm(request.POST)
        if form.is_valid():
            prosthesis = form.save()
            
            # Automatyczne dopasowanie nowej protezy do wszystkich pacjentów
            for patient in Patient.objects.all():
                score = calculate_match_score(patient, prosthesis)
                PatientProsthesis.objects.create(patient=patient, prosthesis=prosthesis, match_score=score)

            messages.success(request, "Proteza zapisana poprawnie! System zaktualizował profile pacjentów.")
            return redirect("/")
    else:
        form = ProsthesisForm()
    return render(request, "add_prosthesis.html", {"form": form})


def add_company(request):
    if request.method == "POST":
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Firma zapisana poprawnie!")
            return redirect("/")
    else:
        form = CompanyForm()
    return render(request, "add_company.html", {"form": form})


def add_match(request):
    # Ten widok służy już tylko do ręcznych poprawek (np. notatek lekarza), 
    # ponieważ relacje generują się automatycznie.
    if request.method == "POST":
        form = PatientProsthesisForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/patients/")
    else:
        form = PatientProsthesisForm()
    return render(request, "add_match.html", {"form": form})


# ── 1. EKSPORT CSV / XLSX ─────────────────────────────────────────────────────

def export_data(request):
    datasets_info = [
        {"label": "Pacjenci",     "value": "patients",   "icon": "fa-solid fa-users",  "color": "text-primary",
         "desc": "Lista pacjentów wraz z parametrami medycznymi."},
        {"label": "Protezy",      "value": "prostheses", "icon": "fa-solid fa-list",   "color": "text-success",
         "desc": "Katalog protez z wymiarami i cenami."},
        {"label": "Dopasowania",  "value": "matches",    "icon": "fa-solid fa-link",   "color": "text-info",
         "desc": "Przypisania protez do pacjentów wraz z wynikami."},
    ]
    return render(request, "export_data.html", {
        "datasets_info": datasets_info,
        "patients": Patient.objects.all()[:20],
    })


def export_csv(request):
    dataset = request.GET.get("dataset", "patients")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{dataset}.csv"'
    response.write("\ufeff")  # BOM dla Excela

    writer = csv.writer(response)

    if dataset == "patients":
        writer.writerow(["ID", "Imię", "Nazwisko", "Budżet (zł)", "Kończyna", "Strona", "Poziom amputacji", "Obwód kikuta (cm)", "Rozmiar"])
        for p in Patient.objects.all():
            writer.writerow([p.id, p.first_name, p.last_name, p.budget, p.limb_type, p.side, p.amputation_level, p.circumference, p.size])

    elif dataset == "prostheses":
        writer.writerow(["ID", "Nazwa protezy", "Firma", "Cena (zł)", "Kończyna", "Strona", "Poziom amputacji", "Rozmiar"])
        for p in Prosthesis.objects.select_related("company").all():
            writer.writerow([p.id, p.name, p.company.name, p.price, p.limb_type, p.side, p.amputation_level, p.size])

    elif dataset == "matches":
        writer.writerow(["ID pacjenta", "Pacjent", "Proteza", "Wynik dopasowania (%)", "Notatki lekarza"])
        for m in PatientProsthesis.objects.select_related("patient", "prosthesis").all():
            writer.writerow([
                m.patient.id, str(m.patient), m.prosthesis.name,
                m.match_score if m.match_score is not None else "",
                m.doctor_notes,
            ])

    return response


def export_xlsx(request):
    dataset = request.GET.get("dataset", "patients")
    wb = openpyxl.Workbook()

    def style_header(ws, headers):
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3C5E")
            cell.alignment = Alignment(horizontal="center")

    if dataset == "patients":
        ws = wb.active; ws.title = "Pacjenci"
        style_header(ws, ["ID", "Imię", "Nazwisko", "Budżet (zł)", "Kończyna", "Strona", "Poziom amputacji", "Obwód (cm)", "Rozmiar"])
        for p in Patient.objects.all():
            ws.append([p.id, p.first_name, p.last_name, p.budget, p.limb_type, p.side, p.amputation_level, p.circumference, p.size])

    elif dataset == "prostheses":
        ws = wb.active; ws.title = "Protezy"
        style_header(ws, ["ID", "Nazwa protezy", "Firma", "Cena (zł)", "Kończyna", "Strona", "Poziom amputacji", "Rozmiar"])
        for p in Prosthesis.objects.select_related("company").all():
            ws.append([p.id, p.name, p.company.name, p.price, p.limb_type, p.side, p.amputation_level, p.size])

    elif dataset == "matches":
        ws = wb.active; ws.title = "Dopasowania"
        style_header(ws, ["ID pacjenta", "Pacjent", "Proteza", "Wynik (%)", "Notatki"])
        for m in PatientProsthesis.objects.select_related("patient", "prosthesis").all():
            ws.append([m.patient.id, str(m.patient), m.prosthesis.name, m.match_score, m.doctor_notes])

    elif dataset == "all":
        ws1 = wb.active; ws1.title = "Pacjenci"
        style_header(ws1, ["ID", "Imię", "Nazwisko", "Budżet (zł)", "Kończyna", "Strona", "Poziom amputacji", "Obwód (cm)", "Rozmiar"])
        for p in Patient.objects.all():
            ws1.append([p.id, p.first_name, p.last_name, p.budget, p.limb_type, p.side, p.amputation_level, p.circumference, p.size])

        ws2 = wb.create_sheet("Protezy")
        style_header(ws2, ["ID", "Nazwa protezy", "Firma", "Cena (zł)", "Kończyna", "Strona", "Poziom amputacji", "Rozmiar"])
        for p in Prosthesis.objects.select_related("company").all():
            ws2.append([p.id, p.name, p.company.name, p.price, p.limb_type, p.side, p.amputation_level, p.size])

        ws3 = wb.create_sheet("Dopasowania")
        style_header(ws3, ["ID pacjenta", "Pacjent", "Proteza", "Wynik (%)", "Notatki"])
        for m in PatientProsthesis.objects.select_related("patient", "prosthesis").all():
            ws3.append([m.patient.id, str(m.patient), m.prosthesis.name, m.match_score, m.doctor_notes])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{dataset}.xlsx"'
    return response


# ── 2. DYNAMICZNY WYKRES (matplotlib → PNG) ──────────────────────────────────

def chart_png(request):
    chart_type = request.GET.get("type", "budget")

    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor("#f0f4f8")
    ax.set_facecolor("#f0f4f8")

    BLUE   = "#1a3c5e"
    colors = ["#2d6a9f", "#28a745", "#ffc107", "#17a2b8", "#dc3545", "#6f42c1", "#fd7e14"]

    if chart_type == "budget":
        patients = list(Patient.objects.all())
        if patients:
            names   = [f"{p.first_name}\n{p.last_name}" for p in patients]
            budgets = [p.budget for p in patients]
            palette = (colors * ((len(patients) // len(colors)) + 1))[:len(patients)]
            bars = ax.bar(names, budgets, color=palette, edgecolor="white", linewidth=1.5)
            ax.bar_label(bars, fmt="%.0f zł", padding=4, fontsize=9, color=BLUE, fontweight="bold")
            ax.set_title("Budżety pacjentów", fontsize=14, fontweight="bold", color=BLUE, pad=16)
            ax.set_ylabel("Budżet (zł)", color="#6c757d")
        else:
            ax.text(0.5, 0.5, "Brak danych", ha="center", va="center",
                    transform=ax.transAxes, fontsize=14, color="#6c757d")

    elif chart_type == "companies":
        companies = list(Company.objects.all())
        if companies:
            labels = [c.name for c in companies]
            sizes  = [Prosthesis.objects.filter(company=c).count() for c in companies]
            palette = (colors * ((len(labels) // len(colors)) + 1))[:len(labels)]
            wedges, texts, autotexts = ax.pie(
                sizes, labels=labels, autopct="%1.0f%%",
                colors=palette, startangle=140,
                wedgeprops={"edgecolor": "white", "linewidth": 2},
            )
            for at in autotexts:
                at.set_fontsize(10); at.set_fontweight("bold"); at.set_color("white")
            ax.set_title("Protezy według firm", fontsize=14, fontweight="bold", color=BLUE, pad=16)
        else:
            ax.text(0.5, 0.5, "Brak danych", ha="center", va="center",
                    transform=ax.transAxes, fontsize=14, color="#6c757d")

    elif chart_type == "prices":
        prostheses = list(Prosthesis.objects.select_related("company").order_by("price"))
        if prostheses:
            names  = [p.name for p in prostheses]
            prices = [p.price for p in prostheses]
            bars = ax.barh(names, prices, color=BLUE, edgecolor="white")
            ax.bar_label(bars, fmt="%.0f zł", padding=4, fontsize=9, color=BLUE, fontweight="bold")
            ax.set_title("Ceny protez", fontsize=14, fontweight="bold", color=BLUE, pad=16)
            ax.set_xlabel("Cena (zł)", color="#6c757d")
        else:
            ax.text(0.5, 0.5, "Brak danych", ha="center", va="center",
                    transform=ax.transAxes, fontsize=14, color="#6c757d")

    for spine in ax.spines.values():
        spine.set_edgecolor("#dee2e6")
    ax.tick_params(colors="#6c757d")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)

    return HttpResponse(buf.read(), content_type="image/png")


def charts_view(request):
    return render(request, "charts.html")


# ── 3. UPLOAD PLIKU CSV / XLSX → IMPORT DANYCH ───────────────────────────────
def _parse_upload(uploaded, filename):
    if filename.endswith(".csv"):
        content = uploaded.read().decode("utf-8-sig")
        reader  = csv.reader(io.StringIO(content))
        return [row for row in reader if any(cell.strip() for cell in row)]

    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(c) if c is not None else "" for c in row])
        return rows

    else:
        raise ValueError("Nieobsługiwany format. Akceptowane: .csv, .xlsx")
    
EXPECTED_COLUMNS = {
    "patients":   {"count": 7, "names": ["imię", "nazwisko", "budżet", "kończyna", "strona", "poziom", "obwód"]},
    "prostheses": {"count": 7, "names": ["nazwa", "cena", "firma", "kończyna", "strona", "poziom", "rozmiar"]},
}

def _validate_rows(rows, model_type):
    """Zwraca listę błędów lub pustą listę jeśli OK."""
    errors = []
    expected = EXPECTED_COLUMNS[model_type]

    if len(rows) < 2:
        errors.append("Plik zawiera tylko nagłówek — brak danych do importu.")
        return errors

    header = [str(c).strip().lower() for c in rows[0]]
    if len(header) < expected["count"]:
        errors.append(
            f"Za mało kolumn: znaleziono {len(header)}, "
            f"oczekiwano minimum {expected['count']} ({', '.join(expected['names'])})."
        )
        return errors

    bad_rows = []
    for i, row in enumerate(rows[1:], start=2):
        if len(row) < expected["count"]:
            bad_rows.append(f"wiersz {i}: za mało kolumn ({len(row)})")
            continue
            
        if model_type == "patients":
            budget_val = str(row[2]).replace(",", ".").strip()
            circ_val = str(row[6]).replace(",", ".").strip()
            try:
                float(budget_val)
                if circ_val: float(circ_val)
            except ValueError:
                bad_rows.append(f"wiersz {i}: błąd w formacie liczby (budżet lub obwód)")
        else:
            price_val = str(row[1]).replace(",", ".").strip()
            try: 
                float(price_val)
            except ValueError:
                bad_rows.append(f"wiersz {i}: błąd w formacie liczby (cena)")

    if bad_rows:
        errors.append("Błędy w danych: " + "; ".join(bad_rows[:5]))
        if len(bad_rows) > 5:
            errors.append(f"... i {len(bad_rows) - 5} więcej błędów.")

    return errors


def import_file(request):
    if request.method != "POST":
        return render(request, "import_file.html", {"preview": None})

    uploaded   = request.FILES.get("file")
    model_type = request.POST.get("model_type", "patients")

    if not uploaded:
        messages.error(request, "Nie wybrano pliku.")
        return render(request, "import_file.html", {"preview": None})

    filename = uploaded.name.lower()
    if not filename.endswith((".csv", ".xlsx", ".xls")):
        messages.error(request, f"Nieobsługiwany format pliku '{uploaded.name}'. Akceptowane: .csv, .xlsx")
        return render(request, "import_file.html", {"preview": None})

    if uploaded.size > 10 * 1024 * 1024:
        messages.error(request, "Plik jest za duży. Maksymalny rozmiar to 10 MB.")
        return render(request, "import_file.html", {"preview": None})

    try:
        rows = _parse_upload(uploaded, filename)
    except Exception as e:
        messages.error(request, f"Nie udało się odczytać pliku: {e}")
        return render(request, "import_file.html", {"preview": None})

    if not rows:
        messages.warning(request, "Plik jest pusty.")
        return render(request, "import_file.html", {"preview": None})

    validation_errors = _validate_rows(rows, model_type)
    if validation_errors:
        for err in validation_errors:
            messages.error(request, err)
        return render(request, "import_file.html", {"preview": rows[:5], "model_type": model_type})

    if "preview" in request.POST:
        return render(request, "import_file.html", {
            "preview": rows[:20],
            "model_type": model_type,
        })

    imported, skipped, errors = 0, 0, []

    for i, row in enumerate(rows[1:], start=2):
        try:
            if model_type == "patients":
                c_val = str(row[6]).replace(",", ".").strip()
                circumference = float(c_val) if c_val else None
                level = str(row[5]).strip()
                p_size = calculate_size(level, circumference)

                patient = Patient.objects.create(
                    first_name=str(row[0]).strip(),
                    last_name=str(row[1]).strip(),
                    budget=float(str(row[2]).replace(",", ".").strip()),
                    limb_type=str(row[3]).strip(),
                    side=str(row[4]).strip(),
                    amputation_level=level,
                    circumference=circumference,
                    size=p_size
                )
                
                # Automatyczne generowanie dopasowań po imporcie
                for prosthesis in Prosthesis.objects.all():
                    score = calculate_match_score(patient, prosthesis)
                    PatientProsthesis.objects.create(patient=patient, prosthesis=prosthesis, match_score=score)

            elif model_type == "prostheses":
                company, _ = Company.objects.get_or_create(name=str(row[2]).strip())
                prosthesis = Prosthesis.objects.create(
                    name=str(row[0]).strip(),
                    price=float(str(row[1]).replace(",", ".").strip()),
                    company=company,
                    limb_type=str(row[3]).strip(),
                    side=str(row[4]).strip(),
                    amputation_level=str(row[5]).strip(),
                    size=str(row[6]).strip()
                )
                
                # Automatyczne generowanie dopasowań po imporcie
                for patient in Patient.objects.all():
                    score = calculate_match_score(patient, prosthesis)
                    PatientProsthesis.objects.create(patient=patient, prosthesis=prosthesis, match_score=score)

            imported += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Wiersz {i}: {e}")

    if imported:
        messages.success(request, f"Zaimportowano {imported} rekordów i przeliczono dopasowania.")
    if skipped:
        messages.warning(request, f"Pominięto {skipped} wierszy ze względu na błędy: {'; '.join(errors[:3])}")

    return redirect("import_file")

# ── 4. USUWANIE PACJENTA ───────────────────────────────

def delete_patient(request, id):
    patient = get_object_or_404(Patient, id=id)
    if request.method == "POST":
        patient.delete()
        messages.success(request, f"Pacjent {patient} został usunięty.")
        return redirect("patients")
    return render(request, "confirm_delete.html", {"patient": patient})